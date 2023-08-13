from docx import Document
from openpyxl import Workbook, load_workbook
from faker import Faker
import os
from datetime import datetime
from random import randint
from time import sleep

#Classe funcionário que receberá cada dado de cada funcionário  
class Employee():
    def __init__(self) -> None:
        self.name = None
        self.cpf = None
        self.rg = None
        self.bday = None
        self.function = None
        self.salary = None
        self.working_schedule = None
        

#Cargos disponibilizados pela empresa
cargos = ['Motorista de Entrega', 'Auxiliar de Carga e Descarga', 'Gerente de Logística', 'Coordenador de Qualidade', 'Gerente de Armazenamento', 'Analista de Dados de Logística', 'Supervisor de Segurança', 'Suporte de Atendimento ao Cliente', 'Planejador de Rotas']
#Salários disponibilizados pela empresa
salarios = ['2300,00', '1850,00', '3500,00', '4850,00', '3400,00', '2950,00', '3200,00', '1600,00', '1900,00']
# 2 jornadas de trabalho, para cargos de supervisão e gestão e para funcionários comuns.
jornada_trabalho = ['A jornada de trabalho do empregado será de 40 horas semanais, de segunda-feira a sexta-feira, das 07h às 18h, com 2 horas de intervalo para almoço.',
                    'A jornada de trabalho do empregado será de 40 horas semanais, de segunda-feira a sexta-feira, das 07:30h às 17:00h, com 1h:30 hora e meia de intervalo para almoço.']


# Aqui verifica se o banco de dados com dados fakes para manipulação não existe, caso não ele cria.
if not os.path.exists('banco_de_dados_fake.xlsx'):
          
    faker = Faker('pt_BR')
    wb = Workbook()
    ws = wb.active
    # cada dado desejado que será inserido na planilha
    colunas = ['NOME', 'RG', 'CPF', 'DATA DE NASC.', 'CARGO', 'REMUNERAÇÃO', 'JORNADA DE TRABALHO']
    ws.append(colunas)
    # A quantidade de funcionários fakes que você deseja criar
    numero_funcionarios = int(input("Quantos funcionários fakes você deseja criar? Digite a quantidade:"))

    # Esta lista marca em que posição a lista de cargos tem cargos de gestão ou supervisão
    lista_hora_gestao = [2,3,4,5]

    for i in range(numero_funcionarios):    
        #se caso o numero randomicamente escolhido para função ser de alguma gestão, vai ter uma carga horaria diferente
        escolhe_funcao_salario = randint(0,8)
        if escolhe_funcao_salario in lista_hora_gestao:
            jornada = jornada_trabalho[0]
        else:
            jornada = jornada_trabalho[1]
        # gerando dados fakes
        funcionario = [
            faker.name(),
            faker.rg(),
            faker.cpf(),
            faker.date_of_birth(minimum_age=18, maximum_age=53).strftime('%d/%m/%Y'),
            cargos[escolhe_funcao_salario],
            salarios[escolhe_funcao_salario],
            jornada
        ]
        # inserindo os dados fakes criados na planilha
        ws.append(funcionario)
    #criando e salvando a planilha
    nome_planilha = 'banco_de_dados_fake.xlsx'
    wb.save(nome_planilha)

    print(f'Banco de Dados Gerados com {numero_funcionarios} funcionários fakes!')



sleep(0.5)

############### AGORA COM BANCO DE DADOS PARA MANIPULAÇÃO E AUTOMAÇÃO DE ARQUIVO ##################

# Como agora temos um banco de dados fakes criados, vamos começar manipular o documento, no caso um modelo de contrato de trabalho

# Primeiro criamos a pasta onde serão guardados os termos
nome_pasta = 'CONTRATOS GERADOS'
if not os.path.exists(nome_pasta):
    os.makedirs(nome_pasta)
    print(f'A pasta {nome_pasta} foi criada.')


    

 
#depois abrir um docx de modelo de contrato
wb = load_workbook('banco_de_dados_fake.xlsx')
ws = wb.active

#contar funcionarios, caso o banco de dados já exista
qtd_funcionario = 0

while(True):
    if(ws[f'A{qtd_funcionario+2}']).value != None:
        qtd_funcionario += 1
    else:
        break


#iterar dentro da planilha fake de banco de dados e popular o contrato 
for i in range(2,(qtd_funcionario+2)):
    document = Document('Modelo Contrato de Trabalho.docx')
    employee = Employee()
    employee.name = ws[f'A{i}'].value
    employee.rg = ws[f'B{i}'].value
    employee.cpf = ws[f'C{i}'].value
    employee.bday = ws[f'D{i}'].value
    employee.function = ws[f'E{i}'].value
    employee.salary  = ws[f'F{i}'].value
    employee.working_schedule = ws[f'G{i}'].value

    for paragraph in document.paragraphs:
        paragraph.text = paragraph.text.replace("employee.name", employee.name)
        paragraph.text = paragraph.text.replace("employee.cpf", employee.cpf)
        paragraph.text = paragraph.text.replace("employee.rg", employee.rg)
        paragraph.text = paragraph.text.replace("employee.bday", employee.bday)
        paragraph.text = paragraph.text.replace("employee.function", employee.function)
        paragraph.text = paragraph.text.replace("employee.salary", employee.salary)
        paragraph.text = paragraph.text.replace("employee.working_schedule", employee.working_schedule)

        contrato_funcionario = f'{nome_pasta}\\Contrato - {employee.function} - {employee.name}.docx'

        document.save(contrato_funcionario)
    employee = None
    
print(f"{qtd_funcionario} Contratos de Trabalho Criados com sucesso!")

  